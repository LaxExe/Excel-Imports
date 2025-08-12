import json
import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from address import street_and_postal_code, street_and_city, fill_missing_address_with_geopy
 
def find_last_row_with_data(ws):
  for row in reversed(range(2, ws.max_row + 1)):
    if any(cell.value not in (None, "") for cell in ws[row]):
      return row
  return 1 # First row if no data found
 
def extract_remake_number(filename):
    # Extract number from remake#.json to sort numerically
    match = re.search(r'remake(\d+)\.json', filename)
    return int(match.group(1)) if match else float('inf')

def build_full_address(address_info):
    if not address_info:
        return "Missing"

    street = address_info.get("street_address", "")
    city = address_info.get("city", "")
    postal = address_info.get("postal_code", "")
    province = address_info.get("province_or_state_name", "")
    country = address_info.get("country", "")

    if street and city and postal:
        return f"{street}, {city}, {province}, {country}, {postal}"

    # Geopy fallback logic
    if not city and street and postal:
        return street_and_postal_code(street, postal)
    elif not postal and street and city:
        return street_and_city(street, city)
    elif street and (city or postal):
        return street_and_postal_code(street, postal or "")
    
    return "Missing"



def append_cleaned_json_to_excel(directory, output_excel):
 
  results = []
  # Dictionary to collect discarded rows
  discarded_rows = {}  
  
  # If output_excel doesn't exist, exit the function

  # Otherwise, load the existing workbook
  wb = load_workbook(output_excel)
  ws = wb.active
 
  # grab headers from the first row
  headers = [cell.value for cell in ws[1]] # travels linearly through the first row
 
  # Row thingy
  last_data_row = find_last_row_with_data(ws)
  next_row = last_data_row + 1
 
  # Seperator shade - apply shade periodically via for loop and update the next_row variable
  gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
  for col in range(1, len(headers) + 1):
    ws.cell(row=next_row, column=col).fill = gray_fill
  next_row += 1
 
  # Get and sort remake files numerically
  remake_files = [
    f for f in os.listdir(directory)
    if re.match(r'remake\d+\.json$', f)
  ]
  print("Remake files found:", remake_files)
  sorted_files = sorted(remake_files, key=extract_remake_number)
 
  # Append data from each JSON file in directory
  for file in sorted_files:
    with open(os.path.join(directory, file), "r") as f:
      data = json.load(f)
    print(f"Processing {file}...")

    count = 0
    # Get clean items list
    clean_items = data.get("clean_items", []) # -> Fall back to empty list if not found
    for item in clean_items:
      print("Processing row:", count) 
      count += 1
      shipping_address_info = (item.get("shipping_address") or [{}])[0]
      billing_address_info = (item.get("billing_address") or [{}])[0]

      if shipping_address_info != billing_address_info: 
        
        full_shipping_address = build_full_address(shipping_address_info)
        full_billing_address = build_full_address(billing_address_info)
      
      else:
        full_shipping_address = shipping_address_info.get("full_address") or build_full_address(shipping_address_info)
        full_billing_address = billing_address_info.get("full_address") or build_full_address(billing_address_info)

      # Get the name
      name = item.get("full_name", "value-missing")
      
      #  Check if name is empty and set display_name
      if name == "":
          display_name = ""  
          # TODO: Add color formatting for empty name cells
      else:
          display_name = name

      # email --> name/number --> discard logic
      # Check if we have the required identifiers
      has_email = item.get("email") and item.get("email") not in [None, "None", "value-missing", ""]
      has_name = name and name not in [None, "None", "value-missing", ""]
      has_phone = item.get("phone_number") and item.get("phone_number") not in [None, "None", "value-missing", ""]

      # Must have name AND (email OR phone)
      if not has_name or (not has_email and not has_phone):
          print(f"Row {count} discarded: Missing name or both email and phone")
          
          #  key - row # and value # (collect discarded rows)
          discarded_rows[count] = {
              "email": item.get("email", "value-missing"),
              "phone_number": item.get("phone_number", "value-missing"),
              "full_name": name,
              "shipping_address": full_shipping_address,
              "billing_address": full_billing_address,
              "additional_fields": item.get("additional_fields", "value-missing")
          }
          continue  # Skip this row

      # Append dict (for valid rows)
      # If we get here, the data is GOOD - append to Excel
      row_data = {
          "email": "" if item.get("email") in [None, "None", "value-missing"] else item.get("email", ""),
          "phone_number": "" if item.get("phone_number") in [None, "None", "value-missing"] else item.get("phone_number", ""),
          "full_name": display_name if display_name not in [None, "None", "value-missing", ""] else name,
          "shipping_address": full_shipping_address,
          "billing_address": full_billing_address,
          "additional_fields": item.get("additional_fields", "")
      }

      # Ensure no None values remain in row_data
      for key in row_data:
          if row_data[key] is None:
              row_data[key] = ""

      # Add color formatting for empty name cells
      if display_name == "":
          # Find the "Full Name" column and add red background
          try:
              name_col_index = headers.index("Full Name") + 1
              cell = ws.cell(row=next_row, column=name_col_index)
              cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
          except ValueError:
              pass  # Column not found, skip coloring

      # Append GOOD data to Excel - ensure no None values
      excel_row = []
      for header in headers:
          value = row_data.get(header, "")
          if value is None:
              value = ""
          excel_row.append(value)
      
      ws.append(excel_row)
      results.append(row_data)
      next_row += 1
 
  
  # Save discarded rows to a file for review
  if discarded_rows:
      with open("discarded_rows.json", "w") as f:
          json.dump(discarded_rows, f, indent=2)
      print(f"Discarded rows saved to discarded_rows.json")
  
  wb.save(output_excel)
  print(results)
  
  return results, discarded_rows  # Return both valid and discarded data
 
