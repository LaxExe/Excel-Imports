from openpyxl import load_workbook
import json
from clean import validate_email, filter_full_name, clean_phone_number, AI_check
from address import is_1_column_tag, column_1_address_skip, column_multi_address
from jsons_to_excel import find_last_row_with_data

# Helper function to extract column letter from new structure
def col_letter_to_index(letter):
  if not letter or letter == None:
    return None
  total = 0
  for i, c in enumerate(reversed(letter.upper())):
    total += (ord(c) - 64) * (26 ** i)
  return total


def gather_row_data(excel_file, json_structure):

  sample = 0
  check = 0
  page = 0
  
  wb = load_workbook(excel_file)
  ws = wb.active
  good_data = []
  results = []
  failed_results = []

  last_row = find_last_row_with_data(ws)

  # Helper function to extract column letter from new structure
  def get_column_letter(field_data):
      if field_data is None:
          return None
      if isinstance(field_data, dict) and "column_letter" in field_data:
          return field_data["column_letter"]
      return None

  # Helper function to extract non-column-letter fields from new structure
  def get_field_value(field_data):
      if field_data is None:
          return None
      if isinstance(field_data, dict):
          # For new structure, return the value directly (not column_letter)
          if "column_letter" in field_data:
              # This is a column field, return None for non-column fields
              return None
          else:
              # This is a value field, return the first non-column_letter value
              for key, value in field_data.items():
                  if key != "column_letter" and key != "column_name" and key != "column_example_data":
                      return value
              return None
      return field_data

  # Get required format 
  required_format = get_field_value(json_structure["address_1_column_format"]["ideal_address_format"])

  if is_1_column_tag("info.json"):

    # Handle additional fields with new structure
    additional_fields = json_structure.get("additional_fields", {})

    col_indexes = {
      "email": col_letter_to_index(get_column_letter(json_structure["required_fields"]["email"])),
      "phone_number": col_letter_to_index(get_column_letter(json_structure["required_fields"]["phone_number"])),
      "full_name": col_letter_to_index(get_column_letter(json_structure["required_fields"]["full_name"])),
      "last_name": col_letter_to_index(get_column_letter(json_structure["required_fields"]["last_name"])),
      "address" : col_letter_to_index(get_column_letter(json_structure["address_1_column_format"]["address_column"])),
      "shipping_address": col_letter_to_index(get_column_letter(json_structure["separate_shipping_and_billing_addresses"]["shipping_address_column"])),
      "billing_address": col_letter_to_index(get_column_letter(json_structure["separate_shipping_and_billing_addresses"]["billing_address_column"]))
    }

    # Handle additional fields with new structure
    for key, field_data in additional_fields.items():
        if field_data is not None:  # Check if field exists
            col_letter = get_column_letter(field_data)
            if col_letter:
                col_indexes[key] = col_letter_to_index(col_letter)

    address_1_column_format = get_field_value(json_structure["address_1_column_format"]["address_format"])
    address_1_column_seperator = get_field_value(json_structure["address_1_column_format"]["address_separator"])

    # Iterates through each data row and extracts the basic information (email, phone, name).
    for i in range(2, last_row + 1):
    
      email = ws.cell(row=i, column=col_indexes["email"]).value
      phone_number = ws.cell(row=i, column=col_indexes["phone_number"]).value
      full_name = ws.cell(row=i, column=col_indexes["full_name"]).value

      # Case 1: Last Name Provided but None
      if (col_indexes["last_name"] == None):
        validate_name = filter_full_name(full_name, email, col_indexes["last_name"])  
      else:
        last_name = ws.cell(row=i, column=col_indexes["last_name"]).value
        validate_name = filter_full_name(full_name, email, last_name)


      # Get Additional Fields
      items = []
      for key in additional_fields:
        if additional_fields != None:
          items.append(key + " : "+ (ws.cell(row=i, column=col_indexes[key])).value + "   ")
      valid_items = ' '.join(items)

      # Validate Email and Phone Number
      valid_email = validate_email(email)
      valid_phone_number = clean_phone_number(phone_number, validate_name, valid_email, None) # Enter the country at none 
      
      # Get adddress field
      # 1. If address exists as a single column, use that value for both shipping and billing address
      # 2. If both shipping and billing address exist, use those values instead

      # Address == billing address
      address = ws.cell(row=i, column=col_indexes["address"]).value
      valid_address = column_1_address_skip(address, address_1_column_format, address_1_column_seperator, required_format)       

      shipping_col = col_indexes.get("shipping_address")
      billing_col = col_indexes.get("billing_address")
      
      # Defaults
      shipping_address = address
      billing_address = address

      if shipping_col is not None and billing_col is not None:
        # Override the default case if both values exists
        shipping_val = ws.cell(row=i, column=shipping_col).value
        billing_val = ws.cell(row=i, column=billing_col).value
        
        # If both values are distinct
        if shipping_val and billing_val and shipping_val != billing_val:
          # Update the default case
          shipping_address = shipping_val
          billing_address = billing_val
        else:
          # Fall back to defualt case + exception if one exists and the other doesn't
          shipping_address = address or shipping_val
          billing_address = address or billing_val
      
      # At this point, either the single address is updated to two new columns or both shipping and billing address are considered
      validate_shipping_address = column_1_address_skip(shipping_address, address_1_column_format, address_1_column_seperator, required_format)
      validate_billing_address = column_1_address_skip(billing_address, address_1_column_format, address_1_column_seperator, required_format)
      
      # Validate Shipping and Billing Address
      if (validate_shipping_address == True or validate_billing_address or valid_email == "Missing" or valid_phone_number == False  or validate_name == False):
        failed_results.append (   
        f'{{\n  "email": "{email}",\n  "phone_number": {valid_phone_number},\n  "full_name": "{validate_name}",\n  "shipping_address": "{shipping_address}",\n "billing_address": "{billing_address}",\n "additional_fields": "{valid_items}"\n}}')
        check = check + 1


      # Validate Address
      if valid_address != True:
        if sample < 3:
          good_data.append (   
          f'{{\n  "email": "{email}",\n  "phone_number": {valid_phone_number},\n  "full_name": "{validate_name}",\n  "shipping_address": "{shipping_address}",\n "billing_address": "{billing_address}",\n "additional_fields": "{valid_items}"\n}}'
          )
          sample = sample + 1


        results.append({
          "email": valid_email,
          "phone_number": valid_phone_number,
          "full_name": validate_name,
          "shipping_address": shipping_address,
          "billing_address": billing_address,
          "additional_fields" : valid_items
        })
  
      if check == 10:
        AI_check(good_data, failed_results, page)
        failed_results = []
        check = 0 
        page = page + 1
  
  else:
    # CASE 2: 
    # This handles Excel files where address information is spread across multiple columns
    # Instead of one "Address" column, we have separate columns for street, city, province, etc.
    
      for i in range(2, last_row + 1):
        # Create a dictionary to hold all the cell values for this row
        # Format: {"A2": "john@email.com", "B2": "123 Main St", "C2": "Toronto", ...}
        row_dict = {}
        
        # Iterate through each column in this specific row
        # This builds our row_dict with all the cell values for processing
        for col in ws.iter_cols(min_row=i, max_row=i):
          cell = col[0]  # Get the cell from this column at row i
          # Store the cell value with its reference (e.g., "A2", "B2") as the key
          # Convert to string and strip whitespace, default to empty string if cell is None
          row_dict[f"{cell.column_letter}{cell.row}"] = str(cell.value).strip() if cell.value else ""

        # Helper function to extract column letter from new structure
        def get_column_letter(field_data):
            if field_data is None:
                return None
            if isinstance(field_data, dict) and "column_letter" in field_data:
                return field_data["column_letter"]
            elif isinstance(field_data, str):
                return field_data  # Handle old format for backward compatibility
            return None

        # Extract the required fields using the JSON structure mapping
        # The JSON tells us which columns contain email, phone, name, etc.
        email_col = get_column_letter(json_structure['required_fields']['email'])
        phone_col = get_column_letter(json_structure['required_fields']['phone_number'])
        full_name_col = get_column_letter(json_structure['required_fields']['full_name'])
        last_name_col = get_column_letter(json_structure['required_fields']['last_name'])
        
        email = row_dict.get(f"{email_col}{i}", "") if email_col else ""
        phone_number = row_dict.get(f"{phone_col}{i}", "") if phone_col else ""
        full_name = row_dict.get(f"{full_name_col}{i}", "") if full_name_col else ""
        
        # Handle last name field - it might not exist in the JSON structure
        # If last_name is null in JSON, we pass None to the name validation function
        last_name_val = None
        if last_name_col:
            last_name_val = row_dict.get(f"{last_name_col}{i}", "")
        
        # Validate and format the full name using the name cleaning function
        # This handles cases where we have separate first/last name columns
        validate_name = filter_full_name(full_name, email, last_name_val)

        # Validate the email address using the email cleaning function
        # This ensures the email is properly formatted and valid
        valid_email = validate_email(email)

        # Clean and format the address using the multi-column address function
        # This is the key step for multi-column addresses - it combines separate columns into one address
        address_result = column_multi_address(row_dict, i, json_path="info.json")
        formatted_address = address_result["address"]  # The complete formatted address string
        needs_ai = address_result["needs_ai"]          # Which address parts need AI to fill in

        # Use the formatted address for phone number validation
        # If the address is complete (no AI needed), we can use it to help validate the phone number
        # If address needs AI, we pass None so phone validation doesn't rely on incomplete address
        address_for_phone = formatted_address if not needs_ai else None
        valid_phone_number = clean_phone_number(phone_number, validate_name, valid_email, address_for_phone)

        # Process additional fields (any columns that aren't email, phone, name, or address)
        # These might be things like "Unit ID", "Owner", "Department", etc.
        items = []
        additional_fields = json_structure.get("additional_fields", {})

        # Loop through each additional field and get its value from the row
        for key, field_data in additional_fields.items():
            if field_data is not None:  # Check if field exists
                col_letter = get_column_letter(field_data)
                if col_letter:
                    cell_key = f"{col_letter}{i}"  
                    val = row_dict.get(cell_key, "")  
                    items.append(f"{key}: {val}")
        
        # Join all additional fields with triple spaces for readability
        # This creates a string like "Unit ID: asdf123   Owner: Nice"
        valid_items = '   '.join(items)

        # Determine if this row needs to be sent to AI for processing
        # A row needs AI if any of these conditions are true:
        # - Email is invalid
        # - Phone number is invalid  
        # - Name is invalid
        # - Address needs AI (missing or invalid address parts)

        results.append({
          "email": valid_email,
          "phone_number": valid_phone_number,
          "full_name": validate_name,
          "address": formatted_address,
          "additional_fields" : valid_items
        })
  

    # Send the good examples and failed results to AI for processing
    # The AI will try to fix any issues in the failed results using the good examples as reference
  AI_check(good_data, failed_results, page)
  return results




