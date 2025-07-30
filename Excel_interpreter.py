# Takes in an Excel file
# Reads the first 5 rows of data, and runs a loop to find last colum needed
# Create a new Excel file with the data

# Feed Data into OpenAI API
# via prompt we will ask it for informatin about the data
    # Identify the required columns
        # Phone Number
        # Email
        # Full Name
        # Adress
            # Adress Format
                # Adress Information
        # Identify Feilds with additional Information

# Generate Json Object with the following feilds

from openpyxl import load_workbook, Workbook
from openai import OpenAI
from dotenv import load_dotenv
from clean import clean_ai_response
import os
import json



load_dotenv()
api_key = os.getenv("OPENAI_API_KEY")

client = OpenAI(
    api_key=api_key,
    base_url="https://api.sambanova.ai/v1"
)



def get_first_5_rows_as_dict(file_path):
    
    wb = load_workbook(file_path)
    ws = wb.active
    cell_dict = {}

    def get_column_letter(col_num):
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result

    row_index = 1
    for row in ws.iter_rows(values_only=True, max_row=5):
        col_index = 1
        for cell_value in row:
            if cell_value is None or str(cell_value).strip() == "":
                remaining_cells = row[col_index:]
                if not any(cell is not None and str(cell).strip() != "" for cell in remaining_cells):
                    break
            
            column_letter = get_column_letter(col_index)
            cell_reference = f"{column_letter}{row_index}"
            
            cell_dict[cell_reference] = str(cell_value) if cell_value is not None else ""
            
            col_index += 1
        row_index += 1

    data_string = json.dumps(cell_dict)
    return data_string

primary_excel_prompt = """
You will be provided with Excel data in dictionary format, where keys are cell references (like "A1", "B2") and values are the cell contents.

Your task is to analyze the header row (row 1) and sample data to identify column mappings and address format.

Step-by-step process:

1. **Extract the header row**: Look at all keys ending with "1" (A1, B1, C1, etc.) to get column names and their corresponding column letters.

2. **Identify required field columns** by matching header names using flexible, case-insensitive patterns:
   - **Phone**: Headers containing words like "phone", "tel", "telephone", "mobile", "cell", "contact number"
   - **Email**: Headers containing words like "email", "e-mail", "mail", or containing "@" symbol
   - **Full Name**: Headers containing words like "name", "full name", "contact person", "client name", "customer name" (exclude business/company names)
   - **Last Name**: Headers containing "last name", "surname", "family name" (optional, can be null if not found)

3. **Analyze address structure**:
   - **Single column**: Look for headers like "address", "full address", "complete address", "location", "mailing address"
   - **Multiple columns**: Look for separate headers like "street", "street address", "address line", "city", "town", "state", "province", "region", "postal code", "zip code", "zip", "country"

4. **Determine address format**:
   - If single column found: Set `address_takes_up_1_column` to true
   - If multiple address columns found: Set `address_takes_up_1_column` to false
   - If both exist, prioritize the structure that has more complete information

5. **For single-column addresses - DETAILED COMPONENT ANALYSIS**:
   - Find the single address column and examine 3-4 actual address values from rows 2-5
   - Indicate the column the address is in.
   - **Identify the separator**: Look at the actual data to find what character separates components (comma, semicolon, pipe, etc.)
   - **Split each address**: Break down each sample address using the identified separator
   - **Component identification rules**:
     - **Street**: Usually contains numbers followed by street names, may include apartment/unit numbers
     - **City**: Typically text-only names without numbers or postal code patterns
     - **State/Province**: Usually 2-3 character codes or full state/province names
     - **Postal Code**: Contains specific patterns:
       - Canadian: Letter-Number-Letter Number-Letter-Number format
       - US: 5 digits or 5+4 format (##### or #####-####)
       - Other countries: Various numeric or alphanumeric patterns
     - **Country**: Full country names or standard country codes
   
   - **Sequential component analysis**: For each address component position:
     - First component: Almost always street address
     - Middle components: Analyze based on content patterns
     - Last component: Often postal code (check for postal patterns first)
     - Second-to-last component: Often state/province if last is postal code
   
   - **Pattern-based identification**: 
     - If a component matches postal code patterns, classify as "postal"
     - If a component is 2-3 characters and appears before postal code, classify as "state"
     - If a component is text-only and appears between street and state/postal, classify as "city"
     - If a component contains street indicators (St, Ave, Blvd, Rd, etc.) or starts with numbers, classify as "street"
   
   - **Only include components that exist**: Count the actual separated components and describe only those that appear
   - **No artificial grouping**: If components are clearly separated by the main separator, treat them as separate components

6. **For multi-column addresses**:
   - Map each address component to its column letter using flexible matching:
     - **street_address**: "street", "address", "address line", "street address"
     - **city**: "city", "town", "municipality"
     - **postal**: "postal", "zip", "postal code", "zip code", "postcode"
     - **province**: "state", "province", "region", "state/province"
     - **country**: "country", "nation"
   - Use "null" if a component cannot be found

7. **Additional fields**:
   - Include any remaining columns that weren't mapped to name, phone, email, or address components
   - Use the exact header name as it appears in the data

**CRITICAL RULES**:
- Examine actual data values, not assumptions
- Use pattern recognition for postal codes and state/province codes
- Respect the natural separation of components by the main separator
- Count components accurately based on actual splits
- Identify each component based on its content pattern, not just position

**Matching Guidelines**:
- Use partial, case-insensitive string matching
- Handle common variations and abbreviations
- Prioritize more specific matches over generic ones
- If multiple columns could match the same field, choose the most specific or complete one

Return ONLY this JSON structure with no additional text:

"""


primary_json_structure = """
{
  "required_fields": {
    "phone_number": "Column Letter",
    "email": "Column Letter",
    "full_name": "Column Letter",
    "last_name": "Column Letter or null",
    "address_takes_up_1_column": true
  },
  "address_1_column_format": {
    "address_column": "Column Letter or null",
    "address_format": "",
    "address_separator": ""
  },
  "if_multi_column_address": {
    "street_address": "Column Letter or null",
    "city": "Column Letter or null",
    "postal": "Column Letter or null"
    "province_or_statename": this feild is optional only include it when province or statename is present
    "country" : this feild is optional only include it when present
  },
  "additional_address_information": {
    "province": "Column Letter or null",
    "country": "Column Letter or null"
  },
  "additional_fields": {
    "Column Name 1": "Column Letter",
    "Column Name 2": "Column Letter"
  }
}
"""





def AI_generate_json_structure(data):


    response = client.chat.completions.create(
                model="Llama-3.3-Swallow-70B-Instruct-v0.4",
                messages=[
                    {
                        "role": "user",
                        "content": primary_excel_prompt + primary_json_structure + data
                    }
                ]
            )
    
    response_content = response.choices[0].message.content.strip()

    print(response_content)

    try:
        response_content = clean_ai_response(response_content)
        info_json = json.loads(response_content)
        required_fields = ["phone_number", "email", "full_name", "address_takes_up_1_column"]
        all_required_present = True

        required_data = info_json.get("required_fields", {})



        for field in required_fields:
            value = required_data.get(field)
            if value in [None, "", "null"]:
                print(f"Missing required field '{field}' in the excel file.")
                print(all_required_present)
                break

        if all_required_present:
            with open(f"info.json", "w") as f:
                json.dump(info_json, f, indent=4) 


    except json.JSONDecodeError:
        print("Error decoding JSON response from OpenAI API.")
        pass

    return "info.json"


































